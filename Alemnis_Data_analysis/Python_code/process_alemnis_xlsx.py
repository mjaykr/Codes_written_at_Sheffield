#!/usr/bin/env python3
"""
Batch-process all Alemnis .xlsx files located beside this Python script.

For every workbook, three independent output sets can be created:

1. Raw processed output
   - Removes malformed/initial time-reset rows.
   - Converts corrected displacement to micrometres.
   - Converts corrected load to millinewtons.
   - Synchronizes load and displacement to one time axis.
   - Preserves the full approach/contact/loading history.

2. Contact-corrected output
   - Detects a statistically significant, sustained load pickup.
   - Backtracks from the confirmed pickup to the estimated response onset.
   - Removes all earlier approach/slip motion.
   - Sets the detected contact point to:
         time = 0 s
         displacement = 0 micrometres
         load = 0 mN

3. Endpoint-trimmed output
   - Is generated only after the first two output sets are complete.
   - Requests an approximate terminal displacement from the user, unless the
     value is supplied through the command line.
   - Searches after peak load for the nearby local load minimum.
   - Treats that minimum as the experimental endpoint.
   - Retains the contact-corrected data from contact through that endpoint.

Plots use SciencePlots' "science" and "ieee" styles with LaTeX enabled
by default.

Default output folders:
    raw_processed_results/<workbook_name>/
    contact_corrected_results/<workbook_name>/
    endpoint_trimmed_results/<workbook_name>/

Each folder contains:
    - one CSV with time, displacement, and load
    - displacement versus time, as PDF and 600 dpi PNG
    - load versus time, as PDF and 600 dpi PNG
    - load versus displacement, as PDF and 600 dpi PNG

The contact-corrected folder also contains a contact-detection diagnostic plot.
The endpoint-trimmed folder contains an endpoint-selection diagnostic plot.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import scienceplots  # noqa: F401  # Registers SciencePlots styles.
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


PREFIX_FACTORS = {
    "f": 1.0e-15,
    "p": 1.0e-12,
    "n": 1.0e-9,
    "u": 1.0e-6,
    "m": 1.0e-3,
}

ZERO_DISPLACEMENT_TOLERANCE_M = 1.0e-15
TIME_MATCH_ABS_TOLERANCE_S = 1.0e-6
MINIMUM_COMPLETE_ROWS = 3


@dataclass(frozen=True)
class ProcessedData:
    time_s: np.ndarray
    displacement_um: np.ndarray
    load_mn: np.ndarray
    removed_initial_rows: int
    interpolated_load: bool
    sheet_name: str


@dataclass(frozen=True)
class ContactDetection:
    contact_index: int
    confirmation_index: int
    loading_direction: int
    baseline_load_mn: float
    baseline_sigma_mn: float
    confirmation_threshold_mn: float
    onset_threshold_mn: float
    smoothed_load_mn: np.ndarray
    smoothing_points: int
    confirmation_points: int


@dataclass(frozen=True)
class ContactCorrectedData:
    time_s: np.ndarray
    displacement_um: np.ndarray
    load_mn: np.ndarray
    contact_time_s: float
    contact_displacement_um: float
    contact_load_mn: float
    detection: ContactDetection


@dataclass(frozen=True)
class EndpointDetection:
    approximate_displacement_um: float
    search_window_um: float
    anchor_index: int
    endpoint_index: int
    peak_index: int
    smoothed_load_mn: np.ndarray
    smoothing_points: int


@dataclass(frozen=True)
class EndpointTrimmedData:
    time_s: np.ndarray
    displacement_um: np.ndarray
    load_mn: np.ndarray
    endpoint_time_s: float
    endpoint_displacement_um: float
    endpoint_load_mn: float
    detection: EndpointDetection


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create raw processed and automatically contact-corrected plots "
            "for every Alemnis .xlsx file beside this script."
        )
    )

    parser.add_argument(
        "--raw-output-dir",
        default="raw_processed_results",
        help=(
            "Raw processed output directory. Relative paths are resolved "
            "beside the script (default: raw_processed_results)."
        ),
    )
    parser.add_argument(
        "--corrected-output-dir",
        default="contact_corrected_results",
        help=(
            "Contact-corrected output directory. Relative paths are resolved "
            "beside the script (default: contact_corrected_results)."
        ),
    )
    parser.add_argument(
        "--endpoint-output-dir",
        default="endpoint_trimmed_results",
        help=(
            "Endpoint-trimmed output directory. Relative paths are resolved "
            "beside the script (default: endpoint_trimmed_results)."
        ),
    )
    parser.add_argument(
        "--end-displacement-um",
        type=float,
        default=None,
        help=(
            "Approximate terminal displacement on the contact-corrected axis, "
            "in micrometres. When omitted, the program prompts separately for "
            "each successfully corrected workbook after the first two output "
            "folders have been generated."
        ),
    )
    parser.add_argument(
        "--endpoint-search-window-um",
        type=float,
        default=0.030,
        help=(
            "Half-width of the displacement interval searched for the nearby "
            "terminal load minimum (default: 0.030 micrometres)."
        ),
    )
    parser.add_argument(
        "--endpoint-smooth-window-s",
        type=float,
        default=0.25,
        help=(
            "Rolling-median window used only to identify the nearby endpoint "
            "minimum (default: 0.25 s)."
        ),
    )
    parser.add_argument(
        "--no-endpoint-prompt",
        action="store_true",
        help=(
            "Do not request endpoint displacement interactively. If "
            "--end-displacement-um is also omitted, the third output branch "
            "is skipped."
        ),
    )
    parser.add_argument(
        "--baseline-duration-s",
        type=float,
        default=5.0,
        help=(
            "Duration used to estimate the pre-contact load baseline "
            "(default: 5.0 s)."
        ),
    )
    parser.add_argument(
        "--smooth-window-s",
        type=float,
        default=0.25,
        help=(
            "Rolling-median window used for contact detection "
            "(default: 0.25 s)."
        ),
    )
    parser.add_argument(
        "--confirmation-duration-s",
        type=float,
        default=0.50,
        help=(
            "Required duration of sustained load pickup "
            "(default: 0.50 s)."
        ),
    )
    parser.add_argument(
        "--sigma-multiplier",
        type=float,
        default=8.0,
        help=(
            "Noise-sigma multiplier for confirming significant load pickup "
            "(default: 8.0)."
        ),
    )
    parser.add_argument(
        "--minimum-pickup-mn",
        type=float,
        default=0.020,
        help=(
            "Minimum load increase/decrease required for confirmation, in mN "
            "(default: 0.020 mN)."
        ),
    )
    parser.add_argument(
        "--onset-sigma-multiplier",
        type=float,
        default=3.0,
        help=(
            "Noise-sigma multiplier used while backtracking to the estimated "
            "response onset (default: 3.0)."
        ),
    )
    parser.add_argument(
        "--minimum-onset-mn",
        type=float,
        default=0.005,
        help=(
            "Minimum load departure used for onset backtracking, in mN "
            "(default: 0.005 mN)."
        ),
    )
    parser.add_argument(
        "--onset-fraction",
        type=float,
        default=0.25,
        help=(
            "Onset threshold as a fraction of the confirmation threshold "
            "above baseline (default: 0.25)."
        ),
    )
    parser.add_argument(
        "--png-dpi",
        type=int,
        default=600,
        help="Resolution of PNG figures (default: 600 dpi).",
    )
    parser.add_argument(
        "--no-latex",
        action="store_true",
        help=(
            "Disable LaTeX rendering. LaTeX is enabled by default as requested."
        ),
    )

    args = parser.parse_args()

    positive_parameters = {
        "--baseline-duration-s": args.baseline_duration_s,
        "--smooth-window-s": args.smooth_window_s,
        "--confirmation-duration-s": args.confirmation_duration_s,
        "--sigma-multiplier": args.sigma_multiplier,
        "--minimum-pickup-mn": args.minimum_pickup_mn,
        "--onset-sigma-multiplier": args.onset_sigma_multiplier,
        "--minimum-onset-mn": args.minimum_onset_mn,
        "--endpoint-search-window-um": args.endpoint_search_window_um,
        "--endpoint-smooth-window-s": args.endpoint_smooth_window_s,
        "--png-dpi": args.png_dpi,
    }
    for name, value in positive_parameters.items():
        if value <= 0:
            parser.error(f"{name} must be greater than zero.")

    if not 0 < args.onset_fraction < 1:
        parser.error("--onset-fraction must lie strictly between 0 and 1.")

    if (
        args.end_displacement_um is not None
        and not np.isfinite(args.end_displacement_um)
    ):
        parser.error("--end-displacement-um must be finite.")

    return args


def _normalise_text(value: object) -> str:
    return (
        str(value)
        .strip()
        .replace("\u2212", "-")
        .replace("\u00b5", "u")
        .replace("\u03bc", "u")
    )


def _time_object_to_seconds(value: time) -> float:
    return (
        value.hour * 3600.0
        + value.minute * 60.0
        + value.second
        + value.microsecond / 1.0e6
    )


def _looks_like_excel_time_format(number_format: str | None) -> bool:
    if not number_format:
        return False

    fmt = number_format.lower()
    return (
        ":" in fmt
        or "[h]" in fmt
        or "[m]" in fmt
        or "[s]" in fmt
        or ("h" in fmt and "s" in fmt)
    )


def parse_time_cell(cell: Cell) -> float:
    """
    Convert an Alemnis time cell to seconds.

    Supported forms:
    - strings such as "-00:00:06.135"
    - datetime.time
    - datetime.datetime
    - datetime.timedelta
    - Excel day fractions
    - numeric seconds
    """
    value = cell.value

    if value is None or value == "":
        return math.nan

    if isinstance(value, timedelta):
        return value.total_seconds()

    if isinstance(value, datetime):
        return _time_object_to_seconds(value.time())

    if isinstance(value, time):
        return _time_object_to_seconds(value)

    if isinstance(value, (int, float, np.integer, np.floating)):
        numeric = float(value)

        if _looks_like_excel_time_format(cell.number_format) or abs(numeric) < 1.0:
            return numeric * 86400.0

        return numeric

    text = _normalise_text(value)
    if not text:
        return math.nan

    try:
        numeric = float(text)
    except ValueError:
        numeric = None

    if numeric is not None:
        if _looks_like_excel_time_format(cell.number_format) or abs(numeric) < 1.0:
            return numeric * 86400.0
        return numeric

    sign = -1.0 if text.startswith("-") else 1.0
    unsigned = text.lstrip("+-").strip()
    parts = unsigned.split(":")

    try:
        if len(parts) == 3:
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = float(parts[2])
            return sign * (hours * 3600.0 + minutes * 60.0 + seconds)

        if len(parts) == 2:
            minutes = float(parts[0])
            seconds = float(parts[1])
            return sign * (minutes * 60.0 + seconds)
    except ValueError as exc:
        raise ValueError(
            f"Cannot parse time value {value!r} in cell {cell.coordinate}."
        ) from exc

    raise ValueError(
        f"Unsupported time value {value!r} in cell {cell.coordinate}."
    )


def parse_measurement_cell(cell: Cell) -> float:
    """
    Convert a displacement/load cell to SI units.

    Numeric cells are assumed to already use SI units:
    - displacement in metres
    - load in newtons

    String suffixes are interpreted as engineering prefixes:
    f, p, n, u/µ/μ, and m.

    Examples:
        655n   -> 655e-9
        1.42m  -> 1.42e-3
        0.5um  -> 0.5e-6
        2.1mN  -> 2.1e-3
    """
    value = cell.value

    if value is None or value == "":
        return math.nan

    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)

    text = _normalise_text(value)
    if not text:
        return math.nan

    match = re.fullmatch(
        r"([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)\s*([A-Za-z]*)",
        text,
    )
    if match is None:
        raise ValueError(
            f"Cannot parse measurement value {value!r} in cell {cell.coordinate}."
        )

    number = float(match.group(1))
    suffix = match.group(2)

    if not suffix:
        return number

    if suffix in PREFIX_FACTORS:
        return number * PREFIX_FACTORS[suffix]

    if suffix in {"N", "M", "meter", "metre"}:
        return number

    if len(suffix) == 2 and suffix[0] in PREFIX_FACTORS and suffix[1] in {"m", "N"}:
        return number * PREFIX_FACTORS[suffix[0]]

    raise ValueError(
        f"Unsupported measurement suffix {suffix!r} in cell {cell.coordinate}."
    )


def choose_data_sheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 2 and worksheet.max_column >= 8:
            return worksheet

    raise ValueError(
        "No worksheet containing at least 8 columns and 2 rows was found."
    )


def read_corrected_channels(xlsx_path: Path) -> tuple[np.ndarray, str]:
    """
    Read corrected channels from workbook columns 1-4.

    Column 1: corrected displacement time
    Column 2: corrected displacement
    Column 3: corrected load time
    Column 4: corrected load
    """
    workbook = load_workbook(
        filename=xlsx_path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = choose_data_sheet(workbook)
        rows: list[list[float]] = []

        for row in worksheet.iter_rows(
            min_row=2,
            max_col=8,
            values_only=False,
        ):
            if all(cell.value in (None, "") for cell in row):
                continue

            rows.append(
                [
                    parse_time_cell(row[0]),
                    parse_measurement_cell(row[1]),
                    parse_time_cell(row[2]),
                    parse_measurement_cell(row[3]),
                ]
            )

        if not rows:
            raise ValueError("The worksheet contains no data rows.")

        return np.asarray(rows, dtype=float), worksheet.title
    finally:
        workbook.close()


def _last_time_reset_start(time_values: np.ndarray) -> int:
    """
    Return the first row of the final strictly increasing time segment.

    Alemnis exports may contain an initial positive timestamp followed by the
    negative pre-test sequence. The final monotonic segment is retained.
    """
    if time_values.size < 2:
        return 0

    resets = np.flatnonzero(np.diff(time_values) <= 0.0)
    return int(resets[-1] + 1) if resets.size else 0


def _find_initial_time_reference(displacement_m: np.ndarray) -> int:
    first = displacement_m[0]

    if first < -ZERO_DISPLACEMENT_TOLERANCE_M:
        return 0

    if abs(first) <= ZERO_DISPLACEMENT_TOLERANCE_M:
        zero_mask = np.abs(displacement_m) <= ZERO_DISPLACEMENT_TOLERANCE_M
        consecutive_end = 0

        for is_zero in zero_mask:
            if is_zero:
                consecutive_end += 1
            else:
                break

        return max(0, consecutive_end - 1)

    return 0


def process_corrected_channels(
    corrected: np.ndarray,
    sheet_name: str,
) -> ProcessedData:
    if corrected.ndim != 2 or corrected.shape[1] != 4:
        raise ValueError("Expected a numeric array with four corrected channels.")

    finite_mask = np.all(np.isfinite(corrected), axis=1)
    data = corrected[finite_mask]

    if data.shape[0] < MINIMUM_COMPLETE_ROWS:
        raise ValueError("Too few complete corrected data rows remain.")

    time_displacement_s = data[:, 0]
    time_load_s = data[:, 2]

    start_index = max(
        _last_time_reset_start(time_displacement_s),
        _last_time_reset_start(time_load_s),
    )

    data = data[start_index:]
    removed_initial_rows = int(np.count_nonzero(~finite_mask) + start_index)

    if data.shape[0] < MINIMUM_COMPLETE_ROWS:
        raise ValueError("Too few rows remain after time-reset correction.")

    time_displacement_s = data[:, 0]
    displacement_m = data[:, 1]
    time_load_s = data[:, 2]
    load_n = data[:, 3]

    if np.any(np.diff(time_displacement_s) <= 0.0):
        raise ValueError("Corrected displacement time is not strictly increasing.")

    if np.any(np.diff(time_load_s) <= 0.0):
        raise ValueError("Corrected load time is not strictly increasing.")

    zero_index = _find_initial_time_reference(displacement_m)
    zero_time_s = float(time_displacement_s[zero_index])

    # One physical time origin is applied to both corrected channels.
    time_displacement_s = time_displacement_s - zero_time_s
    time_load_s = time_load_s - zero_time_s

    displacement_um = displacement_m * 1.0e6
    load_mn_original = load_n * 1.0e3

    times_match = (
        time_displacement_s.shape == time_load_s.shape
        and np.allclose(
            time_displacement_s,
            time_load_s,
            rtol=0.0,
            atol=TIME_MATCH_ABS_TOLERANCE_S,
            equal_nan=False,
        )
    )

    if times_match:
        synchronized_load_mn = load_mn_original
        interpolated_load = False
    else:
        synchronized_load_mn = np.interp(
            time_displacement_s,
            time_load_s,
            load_mn_original,
            left=np.nan,
            right=np.nan,
        )
        interpolated_load = True

    return ProcessedData(
        time_s=time_displacement_s,
        displacement_um=displacement_um,
        load_mn=synchronized_load_mn,
        removed_initial_rows=removed_initial_rows,
        interpolated_load=interpolated_load,
        sheet_name=sheet_name,
    )


def _odd_point_count(duration_s: float, sampling_interval_s: float) -> int:
    points = max(3, int(round(duration_s / sampling_interval_s)))

    if points % 2 == 0:
        points += 1

    return points


def rolling_median(values: np.ndarray, window_points: int) -> np.ndarray:
    if window_points < 3:
        raise ValueError("Rolling-median window must contain at least 3 points.")

    if window_points % 2 == 0:
        window_points += 1

    half_window = window_points // 2
    padded = np.pad(
        values,
        pad_width=(half_window, half_window),
        mode="edge",
    )
    windows = np.lib.stride_tricks.sliding_window_view(
        padded,
        window_shape=window_points,
    )
    return np.median(windows, axis=-1)


def robust_noise_sigma(values: np.ndarray) -> float:
    median = float(np.median(values))
    mad = float(np.median(np.abs(values - median)))
    sigma = 1.4826 * mad

    if not np.isfinite(sigma) or sigma <= 0:
        sigma = float(np.std(values, ddof=1)) if values.size > 1 else 0.0

    if not np.isfinite(sigma) or sigma <= 0:
        sigma = np.finfo(float).eps

    return sigma


def find_first_sustained_run(mask: np.ndarray, required_points: int) -> int | None:
    if required_points <= 0:
        raise ValueError("required_points must be positive.")

    if mask.size < required_points:
        return None

    run_length = 0

    for index, state in enumerate(mask):
        if state:
            run_length += 1
            if run_length >= required_points:
                return index - required_points + 1
        else:
            run_length = 0

    return None


def detect_contact(
    data: ProcessedData,
    *,
    baseline_duration_s: float,
    smooth_window_s: float,
    confirmation_duration_s: float,
    sigma_multiplier: float,
    minimum_pickup_mn: float,
    onset_sigma_multiplier: float,
    minimum_onset_mn: float,
    onset_fraction: float,
) -> ContactDetection:
    finite_mask = (
        np.isfinite(data.time_s)
        & np.isfinite(data.displacement_um)
        & np.isfinite(data.load_mn)
    )

    if np.count_nonzero(finite_mask) < MINIMUM_COMPLETE_ROWS:
        raise ValueError("Insufficient finite data for automatic contact detection.")

    # The synchronized arrays normally contain no internal missing values. If
    # edge interpolation generated NaNs, keep the finite continuous region.
    finite_indices = np.flatnonzero(finite_mask)
    first_finite = int(finite_indices[0])
    last_finite = int(finite_indices[-1])

    if not np.all(finite_mask[first_finite : last_finite + 1]):
        raise ValueError(
            "Internal missing values prevent automatic contact detection."
        )

    time_s = data.time_s[first_finite : last_finite + 1]
    load_mn = data.load_mn[first_finite : last_finite + 1]

    time_steps = np.diff(time_s)
    positive_steps = time_steps[time_steps > 0]

    if positive_steps.size == 0:
        raise ValueError("A positive sampling interval could not be determined.")

    sampling_interval_s = float(np.median(positive_steps))
    smoothing_points = _odd_point_count(
        smooth_window_s,
        sampling_interval_s,
    )
    smoothed_load_mn = rolling_median(load_mn, smoothing_points)

    baseline_end_time = time_s[0] + baseline_duration_s
    baseline_points = int(np.searchsorted(
        time_s,
        baseline_end_time,
        side="right",
    ))
    baseline_points = max(baseline_points, min(20, load_mn.size))
    baseline_points = min(baseline_points, load_mn.size)

    if baseline_points < 3:
        raise ValueError("Too few points are available for baseline estimation.")

    baseline_segment = smoothed_load_mn[:baseline_points]
    baseline_load_mn = float(np.median(baseline_segment))
    baseline_sigma_mn = robust_noise_sigma(baseline_segment)

    positive_excursion = (
        float(np.quantile(smoothed_load_mn, 0.99)) - baseline_load_mn
    )
    negative_excursion = (
        baseline_load_mn - float(np.quantile(smoothed_load_mn, 0.01))
    )
    loading_direction = 1 if positive_excursion >= negative_excursion else -1

    confirmation_delta_mn = max(
        sigma_multiplier * baseline_sigma_mn,
        minimum_pickup_mn,
    )
    onset_delta_mn = max(
        onset_sigma_multiplier * baseline_sigma_mn,
        minimum_onset_mn,
        onset_fraction * confirmation_delta_mn,
    )

    directed_departure_mn = (
        loading_direction * (smoothed_load_mn - baseline_load_mn)
    )

    confirmation_points = max(
        3,
        int(math.ceil(confirmation_duration_s / sampling_interval_s)),
    )

    confirmation_local_index = find_first_sustained_run(
        directed_departure_mn >= confirmation_delta_mn,
        confirmation_points,
    )

    if confirmation_local_index is None:
        maximum_departure = float(np.max(directed_departure_mn))
        raise ValueError(
            "No sustained significant load pickup was detected. "
            f"Maximum directed departure was {maximum_departure:.6g} mN; "
            f"the confirmation threshold was {confirmation_delta_mn:.6g} mN."
        )

    # Backtrack from the confirmed pickup to the last point that was still at
    # or below the lower onset threshold. This estimates the beginning of the
    # material response rather than using the later confirmation crossing.
    prior_at_baseline = np.flatnonzero(
        directed_departure_mn[: confirmation_local_index + 1]
        <= onset_delta_mn
    )

    if prior_at_baseline.size:
        contact_local_index = min(
            int(prior_at_baseline[-1] + 1),
            confirmation_local_index,
        )
    else:
        contact_local_index = 0

    contact_index = first_finite + contact_local_index
    confirmation_index = first_finite + confirmation_local_index

    confirmation_threshold_mn = (
        baseline_load_mn + loading_direction * confirmation_delta_mn
    )
    onset_threshold_mn = (
        baseline_load_mn + loading_direction * onset_delta_mn
    )

    full_smoothed = np.full(data.load_mn.shape, np.nan, dtype=float)
    full_smoothed[first_finite : last_finite + 1] = smoothed_load_mn

    return ContactDetection(
        contact_index=contact_index,
        confirmation_index=confirmation_index,
        loading_direction=loading_direction,
        baseline_load_mn=baseline_load_mn,
        baseline_sigma_mn=baseline_sigma_mn,
        confirmation_threshold_mn=confirmation_threshold_mn,
        onset_threshold_mn=onset_threshold_mn,
        smoothed_load_mn=full_smoothed,
        smoothing_points=smoothing_points,
        confirmation_points=confirmation_points,
    )


def apply_contact_correction(
    data: ProcessedData,
    detection: ContactDetection,
) -> ContactCorrectedData:
    index = detection.contact_index

    contact_time_s = float(data.time_s[index])
    contact_displacement_um = float(data.displacement_um[index])
    contact_load_mn = float(data.load_mn[index])

    time_s = data.time_s[index:] - contact_time_s
    displacement_um = (
        data.displacement_um[index:] - contact_displacement_um
    )
    load_mn = data.load_mn[index:] - contact_load_mn

    # Remove floating-point residuals at the newly defined origin.
    time_s[0] = 0.0
    displacement_um[0] = 0.0
    load_mn[0] = 0.0

    return ContactCorrectedData(
        time_s=time_s,
        displacement_um=displacement_um,
        load_mn=load_mn,
        contact_time_s=contact_time_s,
        contact_displacement_um=contact_displacement_um,
        contact_load_mn=contact_load_mn,
        detection=detection,
    )



def detect_experiment_endpoint(
    data: ContactCorrectedData,
    *,
    approximate_displacement_um: float,
    search_window_um: float,
    smooth_window_s: float,
) -> EndpointDetection:
    """
    Locate the experimental endpoint near a user-supplied displacement.

    The search is performed on the contact-corrected data and is restricted to
    the post-peak portion of the experiment. This prevents an earlier loading
    point with the same displacement from being selected.

    The measured load is smoothed with a rolling median only for endpoint
    detection. The exported and plotted values remain the measured data.
    """
    finite_mask = (
        np.isfinite(data.time_s)
        & np.isfinite(data.displacement_um)
        & np.isfinite(data.load_mn)
    )

    if np.count_nonzero(finite_mask) < MINIMUM_COMPLETE_ROWS:
        raise ValueError("Insufficient finite data for endpoint detection.")

    finite_indices = np.flatnonzero(finite_mask)
    first_finite = int(finite_indices[0])
    last_finite = int(finite_indices[-1])

    if not np.all(finite_mask[first_finite : last_finite + 1]):
        raise ValueError(
            "Internal missing values prevent automatic endpoint detection."
        )

    time_s = data.time_s[first_finite : last_finite + 1]
    displacement_um = data.displacement_um[first_finite : last_finite + 1]
    load_mn = data.load_mn[first_finite : last_finite + 1]

    time_steps = np.diff(time_s)
    positive_steps = time_steps[time_steps > 0]

    if positive_steps.size == 0:
        raise ValueError("A positive sampling interval could not be determined.")

    sampling_interval_s = float(np.median(positive_steps))
    smoothing_points = _odd_point_count(
        smooth_window_s,
        sampling_interval_s,
    )
    smoothed_load_mn = rolling_median(load_mn, smoothing_points)

    loading_direction = data.detection.loading_direction
    directed_smoothed_load = loading_direction * smoothed_load_mn

    peak_local_index = int(np.argmax(directed_smoothed_load))
    post_peak_indices = np.arange(
        peak_local_index,
        displacement_um.size,
        dtype=int,
    )

    if post_peak_indices.size < 2:
        raise ValueError(
            "No usable post-peak data are available for endpoint detection."
        )

    post_peak_displacement = displacement_um[post_peak_indices]
    anchor_local_index = int(
        post_peak_indices[
            np.argmin(
                np.abs(
                    post_peak_displacement
                    - approximate_displacement_um
                )
            )
        ]
    )

    nearby_mask = (
        np.abs(displacement_um - approximate_displacement_um)
        <= search_window_um
    )
    nearby_mask[:peak_local_index] = False

    # Keep only the contiguous nearby-displacement segment containing the
    # nearest post-peak point. This avoids mixing separate passes through the
    # same displacement.
    if nearby_mask[anchor_local_index]:
        segment_start = anchor_local_index
        segment_end = anchor_local_index

        while segment_start > peak_local_index and nearby_mask[segment_start - 1]:
            segment_start -= 1

        while (
            segment_end + 1 < displacement_um.size
            and nearby_mask[segment_end + 1]
        ):
            segment_end += 1

        candidate_indices = np.arange(
            segment_start,
            segment_end + 1,
            dtype=int,
        )
    else:
        candidate_indices = np.asarray([anchor_local_index], dtype=int)

    # Expand a very small candidate region so that a local minimum can be
    # evaluated even when the displacement changes rapidly.
    minimum_candidate_points = max(5, smoothing_points)
    if candidate_indices.size < minimum_candidate_points:
        half_width = minimum_candidate_points // 2
        segment_start = max(
            peak_local_index,
            anchor_local_index - half_width,
        )
        segment_end = min(
            displacement_um.size - 1,
            anchor_local_index + half_width,
        )
        candidate_indices = np.arange(
            segment_start,
            segment_end + 1,
            dtype=int,
        )

    local_minima: list[int] = []
    for index in candidate_indices:
        if index <= peak_local_index or index >= displacement_um.size - 1:
            continue

        if (
            directed_smoothed_load[index]
            <= directed_smoothed_load[index - 1]
            and directed_smoothed_load[index]
            <= directed_smoothed_load[index + 1]
        ):
            local_minima.append(int(index))

    if local_minima:
        # Select the deepest nearby post-peak minimum. Distance from the user's
        # approximate displacement is used only as a tie-breaker.
        endpoint_local_index = min(
            local_minima,
            key=lambda index: (
                directed_smoothed_load[index],
                abs(
                    displacement_um[index]
                    - approximate_displacement_um
                ),
            ),
        )
    else:
        endpoint_local_index = int(
            candidate_indices[
                np.argmin(directed_smoothed_load[candidate_indices])
            ]
        )

    if endpoint_local_index <= peak_local_index:
        raise ValueError(
            "The selected endpoint does not occur after peak load. "
            "Provide a later approximate displacement or enlarge the "
            "endpoint search window."
        )

    full_smoothed = np.full(data.load_mn.shape, np.nan, dtype=float)
    full_smoothed[first_finite : last_finite + 1] = smoothed_load_mn

    return EndpointDetection(
        approximate_displacement_um=float(approximate_displacement_um),
        search_window_um=float(search_window_um),
        anchor_index=first_finite + anchor_local_index,
        endpoint_index=first_finite + endpoint_local_index,
        peak_index=first_finite + peak_local_index,
        smoothed_load_mn=full_smoothed,
        smoothing_points=smoothing_points,
    )


def apply_endpoint_trim(
    data: ContactCorrectedData,
    detection: EndpointDetection,
) -> EndpointTrimmedData:
    endpoint_index = detection.endpoint_index
    retained_slice = slice(0, endpoint_index + 1)

    return EndpointTrimmedData(
        time_s=data.time_s[retained_slice].copy(),
        displacement_um=data.displacement_um[retained_slice].copy(),
        load_mn=data.load_mn[retained_slice].copy(),
        endpoint_time_s=float(data.time_s[endpoint_index]),
        endpoint_displacement_um=float(data.displacement_um[endpoint_index]),
        endpoint_load_mn=float(data.load_mn[endpoint_index]),
        detection=detection,
    )


def prompt_for_endpoint_displacement(
    filename: str,
    data: ContactCorrectedData,
) -> float | None:
    finite_displacement = data.displacement_um[
        np.isfinite(data.displacement_um)
    ]

    minimum_displacement = float(np.min(finite_displacement))
    maximum_displacement = float(np.max(finite_displacement))

    prompt = (
        f"\nApproximate terminal displacement for {filename}, in um "
        f"[contact-corrected range {minimum_displacement:.6g} to "
        f"{maximum_displacement:.6g}; press Enter to skip]: "
    )

    while True:
        try:
            response = input(prompt).strip()
        except EOFError:
            print(
                f"No interactive input is available; endpoint trimming for "
                f"{filename} was skipped."
            )
            return None

        if response == "":
            return None

        response = (
            response
            .replace("\u2212", "-")
            .replace("\u00b5", "")
            .replace("\u03bc", "")
            .strip()
        )

        try:
            value = float(response)
        except ValueError:
            print("Enter a numeric displacement in micrometres, or press Enter.")
            continue

        if not np.isfinite(value):
            print("The displacement must be finite.")
            continue

        return value


def format_number(value: float) -> str:
    if not np.isfinite(value):
        return ""

    if value == 0.0:
        return "0"

    return f"{value:.12g}"


def write_three_column_csv(
    csv_path: Path,
    *,
    headers: tuple[str, str, str],
    time_s: np.ndarray,
    displacement_um: np.ndarray,
    load_mn: np.ndarray,
) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.writer(stream)
        writer.writerow(headers)

        for current_time, displacement, load in zip(
            time_s,
            displacement_um,
            load_mn,
            strict=True,
        ):
            writer.writerow(
                [
                    format_number(float(current_time)),
                    format_number(float(displacement)),
                    format_number(float(load)),
                ]
            )


def configure_plotting(use_latex: bool) -> None:
    plt.rcParams.update(
        {
            "text.usetex": use_latex,
            "font.family": "serif",
            "axes.unicode_minus": False,
        }
    )


def save_plot(
    x: np.ndarray,
    y: np.ndarray,
    *,
    xlabel: str,
    ylabel: str,
    output_stem: Path,
    png_dpi: int,
) -> None:
    finite = np.isfinite(x) & np.isfinite(y)

    if np.count_nonzero(finite) < 2:
        raise ValueError(
            f"Insufficient finite data for plot {output_stem.name!r}."
        )

    output_stem.parent.mkdir(parents=True, exist_ok=True)

    with plt.style.context(["science", "ieee"]):
        figure, axis = plt.subplots(figsize=(3.50, 2.65))

        axis.plot(
            x[finite],
            y[finite],
            linewidth=1.0,
        )
        axis.set_xlabel(xlabel)
        axis.set_ylabel(ylabel)
        axis.tick_params(
            which="both",
            direction="in",
            top=True,
            right=True,
        )
        axis.minorticks_on()
        axis.grid(
            visible=True,
            which="major",
            linewidth=0.35,
            alpha=0.35,
        )
        axis.margins(x=0.02, y=0.04)

        figure.savefig(
            output_stem.with_suffix(".pdf"),
            bbox_inches="tight",
            pad_inches=0.03,
        )
        figure.savefig(
            output_stem.with_suffix(".png"),
            dpi=png_dpi,
            bbox_inches="tight",
            pad_inches=0.03,
        )
        plt.close(figure)


def save_contact_diagnostic(
    data: ProcessedData,
    detection: ContactDetection,
    *,
    output_stem: Path,
    png_dpi: int,
) -> None:
    finite_raw = np.isfinite(data.time_s) & np.isfinite(data.load_mn)
    finite_smoothed = (
        np.isfinite(data.time_s)
        & np.isfinite(detection.smoothed_load_mn)
    )

    output_stem.parent.mkdir(parents=True, exist_ok=True)

    with plt.style.context(["science", "ieee"]):
        figure, axis = plt.subplots(figsize=(3.50, 2.65))

        axis.plot(
            data.time_s[finite_raw],
            data.load_mn[finite_raw],
            linewidth=0.65,
            alpha=0.55,
            label="Measured load",
        )
        axis.plot(
            data.time_s[finite_smoothed],
            detection.smoothed_load_mn[finite_smoothed],
            linewidth=1.0,
            label="Rolling median",
        )

        axis.axhline(
            detection.baseline_load_mn,
            linewidth=0.7,
            linestyle=":",
            label="Baseline",
        )
        axis.axhline(
            detection.onset_threshold_mn,
            linewidth=0.7,
            linestyle="--",
            label="Onset threshold",
        )
        axis.axhline(
            detection.confirmation_threshold_mn,
            linewidth=0.7,
            linestyle="-.",
            label="Confirmation threshold",
        )
        axis.axvline(
            data.time_s[detection.contact_index],
            linewidth=0.8,
            linestyle="--",
            label="Detected contact",
        )

        axis.set_xlabel(r"Time, $t$ (s)")
        axis.set_ylabel(r"Load, $F$ (mN)")
        axis.tick_params(
            which="both",
            direction="in",
            top=True,
            right=True,
        )
        axis.minorticks_on()
        axis.grid(
            visible=True,
            which="major",
            linewidth=0.35,
            alpha=0.35,
        )
        axis.legend(
            loc="best",
            frameon=False,
            fontsize=6,
        )

        figure.savefig(
            output_stem.with_suffix(".pdf"),
            bbox_inches="tight",
            pad_inches=0.03,
        )
        figure.savefig(
            output_stem.with_suffix(".png"),
            dpi=png_dpi,
            bbox_inches="tight",
            pad_inches=0.03,
        )
        plt.close(figure)



def save_endpoint_diagnostic(
    data: ContactCorrectedData,
    detection: EndpointDetection,
    *,
    output_stem: Path,
    png_dpi: int,
) -> None:
    finite_raw = (
        np.isfinite(data.displacement_um)
        & np.isfinite(data.load_mn)
    )
    finite_smoothed = (
        np.isfinite(data.displacement_um)
        & np.isfinite(detection.smoothed_load_mn)
    )

    output_stem.parent.mkdir(parents=True, exist_ok=True)

    with plt.style.context(["science", "ieee"]):
        figure, axis = plt.subplots(figsize=(3.50, 2.65))

        axis.plot(
            data.displacement_um[finite_raw],
            data.load_mn[finite_raw],
            linewidth=0.65,
            alpha=0.55,
            label="Measured load",
        )
        axis.plot(
            data.displacement_um[finite_smoothed],
            detection.smoothed_load_mn[finite_smoothed],
            linewidth=1.0,
            label="Rolling median",
        )

        lower_bound = (
            detection.approximate_displacement_um
            - detection.search_window_um
        )
        upper_bound = (
            detection.approximate_displacement_um
            + detection.search_window_um
        )

        axis.axvspan(
            lower_bound,
            upper_bound,
            alpha=0.12,
            label="Endpoint search interval",
        )
        axis.axvline(
            detection.approximate_displacement_um,
            linewidth=0.75,
            linestyle=":",
            label="Approximate displacement",
        )
        axis.axvline(
            data.displacement_um[detection.endpoint_index],
            linewidth=0.85,
            linestyle="--",
            label="Selected endpoint",
        )
        axis.plot(
            [data.displacement_um[detection.endpoint_index]],
            [data.load_mn[detection.endpoint_index]],
            marker="o",
            linestyle="none",
            markersize=3.5,
        )

        axis.set_xlabel(r"Displacement, $d$ ($\mu\mathrm{m}$)")
        axis.set_ylabel(r"Load, $F$ (mN)")
        axis.tick_params(
            which="both",
            direction="in",
            top=True,
            right=True,
        )
        axis.minorticks_on()
        axis.grid(
            visible=True,
            which="major",
            linewidth=0.35,
            alpha=0.35,
        )
        axis.legend(
            loc="best",
            frameon=False,
            fontsize=6,
        )

        figure.savefig(
            output_stem.with_suffix(".pdf"),
            bbox_inches="tight",
            pad_inches=0.03,
        )
        figure.savefig(
            output_stem.with_suffix(".png"),
            dpi=png_dpi,
            bbox_inches="tight",
            pad_inches=0.03,
        )
        plt.close(figure)


def save_standard_plot_set(
    *,
    output_directory: Path,
    filename_stem: str,
    output_suffix: str,
    time_s: np.ndarray,
    displacement_um: np.ndarray,
    load_mn: np.ndarray,
    png_dpi: int,
) -> None:
    save_plot(
        time_s,
        displacement_um,
        xlabel=r"Time, $t$ (s)",
        ylabel=r"Displacement, $d$ ($\mu\mathrm{m}$)",
        output_stem=(
            output_directory
            / f"{filename_stem}_displacement_vs_time_{output_suffix}"
        ),
        png_dpi=png_dpi,
    )
    save_plot(
        time_s,
        load_mn,
        xlabel=r"Time, $t$ (s)",
        ylabel=r"Load, $F$ (mN)",
        output_stem=(
            output_directory
            / f"{filename_stem}_load_vs_time_{output_suffix}"
        ),
        png_dpi=png_dpi,
    )
    save_plot(
        displacement_um,
        load_mn,
        xlabel=r"Displacement, $d$ ($\mu\mathrm{m}$)",
        ylabel=r"Load, $F$ (mN)",
        output_stem=(
            output_directory
            / f"{filename_stem}_load_vs_displacement_{output_suffix}"
        ),
        png_dpi=png_dpi,
    )


def resolve_output_directory(
    script_directory: Path,
    value: str,
) -> Path:
    path = Path(value).expanduser()

    if not path.is_absolute():
        path = script_directory / path

    return path.resolve()


def write_detection_summary(
    summary_path: Path,
    summary_rows: list[dict[str, object]],
) -> None:
    fieldnames = [
        "Filename",
        "Status",
        "Worksheet",
        "Contact_time_in_raw_s",
        "Contact_displacement_in_raw_um",
        "Contact_load_in_raw_mN",
        "Loading_direction",
        "Baseline_load_mN",
        "Baseline_noise_sigma_mN",
        "Onset_threshold_mN",
        "Confirmation_threshold_mN",
        "Smoothing_points",
        "Confirmation_points",
        "Message",
    ]

    summary_path.parent.mkdir(parents=True, exist_ok=True)

    with summary_path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames)
        writer.writeheader()

        for row in summary_rows:
            writer.writerow(row)



def write_endpoint_summary(
    summary_path: Path,
    summary_rows: list[dict[str, object]],
) -> None:
    fieldnames = [
        "Filename",
        "Status",
        "Approximate_displacement_um",
        "Search_window_um",
        "Peak_time_s",
        "Peak_displacement_um",
        "Peak_load_mN",
        "Selected_endpoint_time_s",
        "Selected_endpoint_displacement_um",
        "Selected_endpoint_load_mN",
        "Retained_rows",
        "Message",
    ]

    summary_path.parent.mkdir(parents=True, exist_ok=True)

    with summary_path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames)
        writer.writeheader()

        for row in summary_rows:
            writer.writerow(row)


def process_workbook(
    xlsx_path: Path,
    *,
    raw_output_root: Path,
    corrected_output_root: Path,
    png_dpi: int,
    contact_parameters: dict[str, float],
) -> tuple[ProcessedData, ContactCorrectedData]:
    corrected_channels, sheet_name = read_corrected_channels(xlsx_path)
    raw_data = process_corrected_channels(
        corrected_channels,
        sheet_name,
    )

    raw_output_directory = raw_output_root / xlsx_path.stem
    corrected_output_directory = (
        corrected_output_root / xlsx_path.stem
    )

    write_three_column_csv(
        raw_output_directory / f"{xlsx_path.stem}_processed_raw.csv",
        headers=("Time_s", "Displacement_um", "Load_mN"),
        time_s=raw_data.time_s,
        displacement_um=raw_data.displacement_um,
        load_mn=raw_data.load_mn,
    )
    save_standard_plot_set(
        output_directory=raw_output_directory,
        filename_stem=xlsx_path.stem,
        output_suffix="raw",
        time_s=raw_data.time_s,
        displacement_um=raw_data.displacement_um,
        load_mn=raw_data.load_mn,
        png_dpi=png_dpi,
    )

    detection = detect_contact(
        raw_data,
        **contact_parameters,
    )
    corrected_data = apply_contact_correction(
        raw_data,
        detection,
    )

    write_three_column_csv(
        corrected_output_directory
        / f"{xlsx_path.stem}_processed_contact_corrected.csv",
        headers=(
            "Time_from_contact_s",
            "Displacement_from_contact_um",
            "Load_from_contact_mN",
        ),
        time_s=corrected_data.time_s,
        displacement_um=corrected_data.displacement_um,
        load_mn=corrected_data.load_mn,
    )
    save_standard_plot_set(
        output_directory=corrected_output_directory,
        filename_stem=xlsx_path.stem,
        output_suffix="contact_corrected",
        time_s=corrected_data.time_s,
        displacement_um=corrected_data.displacement_um,
        load_mn=corrected_data.load_mn,
        png_dpi=png_dpi,
    )
    save_contact_diagnostic(
        raw_data,
        detection,
        output_stem=(
            corrected_output_directory
            / f"{xlsx_path.stem}_contact_detection_diagnostic"
        ),
        png_dpi=png_dpi,
    )

    return raw_data, corrected_data


def main() -> int:
    print("Alemnis processor version: 3.0-three-folder")
    print()
    args = parse_arguments()

    script_directory = Path(__file__).resolve().parent
    raw_output_root = resolve_output_directory(
        script_directory,
        args.raw_output_dir,
    )
    corrected_output_root = resolve_output_directory(
        script_directory,
        args.corrected_output_dir,
    )
    endpoint_output_root = resolve_output_directory(
        script_directory,
        args.endpoint_output_dir,
    )

    xlsx_files = sorted(
        path
        for path in script_directory.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )

    if not xlsx_files:
        print(
            f"No .xlsx files were found beside the script:\n"
            f"{script_directory}",
            file=sys.stderr,
        )
        return 1

    use_latex = not args.no_latex

    if use_latex and shutil.which("latex") is None:
        print(
            "LaTeX rendering is enabled, but the 'latex' executable was not "
            "found on PATH.\n"
            "Install/configure MiKTeX or TeX Live, or use --no-latex only "
            "for diagnostic execution.",
            file=sys.stderr,
        )
        return 1

    configure_plotting(use_latex)
    raw_output_root.mkdir(parents=True, exist_ok=True)
    corrected_output_root.mkdir(parents=True, exist_ok=True)

    contact_parameters = {
        "baseline_duration_s": args.baseline_duration_s,
        "smooth_window_s": args.smooth_window_s,
        "confirmation_duration_s": args.confirmation_duration_s,
        "sigma_multiplier": args.sigma_multiplier,
        "minimum_pickup_mn": args.minimum_pickup_mn,
        "onset_sigma_multiplier": args.onset_sigma_multiplier,
        "minimum_onset_mn": args.minimum_onset_mn,
        "onset_fraction": args.onset_fraction,
    }

    summary_rows: list[dict[str, object]] = []
    corrected_datasets: dict[Path, ContactCorrectedData] = {}
    raw_failures = 0
    correction_failures = 0

    print(f"Input directory           : {script_directory}")
    print(f"Raw output directory      : {raw_output_root}")
    print(f"Corrected output directory: {corrected_output_root}")
    print(f"Endpoint output directory : {endpoint_output_root}")
    print(f"Excel files               : {len(xlsx_files)}")
    print(f"LaTeX rendering           : {use_latex}")
    print()

    # Stage 1 and Stage 2: generate the complete raw-processed and
    # contact-corrected outputs for every workbook.
    for xlsx_path in xlsx_files:
        try:
            corrected_channels, sheet_name = read_corrected_channels(
                xlsx_path
            )
            raw_data = process_corrected_channels(
                corrected_channels,
                sheet_name,
            )

            raw_output_directory = raw_output_root / xlsx_path.stem

            write_three_column_csv(
                raw_output_directory
                / f"{xlsx_path.stem}_processed_raw.csv",
                headers=("Time_s", "Displacement_um", "Load_mN"),
                time_s=raw_data.time_s,
                displacement_um=raw_data.displacement_um,
                load_mn=raw_data.load_mn,
            )
            save_standard_plot_set(
                output_directory=raw_output_directory,
                filename_stem=xlsx_path.stem,
                output_suffix="raw",
                time_s=raw_data.time_s,
                displacement_um=raw_data.displacement_um,
                load_mn=raw_data.load_mn,
                png_dpi=args.png_dpi,
            )

            print(f"[RAW OK] {xlsx_path.name}")

        except Exception as exc:
            raw_failures += 1
            correction_failures += 1
            message = str(exc)
            print(
                f"[RAW FAILED] {xlsx_path.name}: {message}",
                file=sys.stderr,
            )
            summary_rows.append(
                {
                    "Filename": xlsx_path.name,
                    "Status": "raw processing failed",
                    "Worksheet": "",
                    "Message": message,
                }
            )
            continue

        try:
            detection = detect_contact(
                raw_data,
                **contact_parameters,
            )
            corrected_data = apply_contact_correction(
                raw_data,
                detection,
            )

            corrected_output_directory = (
                corrected_output_root / xlsx_path.stem
            )

            write_three_column_csv(
                corrected_output_directory
                / f"{xlsx_path.stem}_processed_contact_corrected.csv",
                headers=(
                    "Time_from_contact_s",
                    "Displacement_from_contact_um",
                    "Load_from_contact_mN",
                ),
                time_s=corrected_data.time_s,
                displacement_um=corrected_data.displacement_um,
                load_mn=corrected_data.load_mn,
            )
            save_standard_plot_set(
                output_directory=corrected_output_directory,
                filename_stem=xlsx_path.stem,
                output_suffix="contact_corrected",
                time_s=corrected_data.time_s,
                displacement_um=corrected_data.displacement_um,
                load_mn=corrected_data.load_mn,
                png_dpi=args.png_dpi,
            )
            save_contact_diagnostic(
                raw_data,
                detection,
                output_stem=(
                    corrected_output_directory
                    / f"{xlsx_path.stem}_contact_detection_diagnostic"
                ),
                png_dpi=args.png_dpi,
            )

            corrected_datasets[xlsx_path] = corrected_data

            direction_text = (
                "positive" if detection.loading_direction > 0 else "negative"
            )

            summary_rows.append(
                {
                    "Filename": xlsx_path.name,
                    "Status": "contact detected",
                    "Worksheet": raw_data.sheet_name,
                    "Contact_time_in_raw_s": format_number(
                        corrected_data.contact_time_s
                    ),
                    "Contact_displacement_in_raw_um": format_number(
                        corrected_data.contact_displacement_um
                    ),
                    "Contact_load_in_raw_mN": format_number(
                        corrected_data.contact_load_mn
                    ),
                    "Loading_direction": direction_text,
                    "Baseline_load_mN": format_number(
                        detection.baseline_load_mn
                    ),
                    "Baseline_noise_sigma_mN": format_number(
                        detection.baseline_sigma_mn
                    ),
                    "Onset_threshold_mN": format_number(
                        detection.onset_threshold_mn
                    ),
                    "Confirmation_threshold_mN": format_number(
                        detection.confirmation_threshold_mn
                    ),
                    "Smoothing_points": detection.smoothing_points,
                    "Confirmation_points": detection.confirmation_points,
                    "Message": "",
                }
            )

            print(
                f"[CORRECTED OK] {xlsx_path.name}: "
                f"contact at t={corrected_data.contact_time_s:.6g} s, "
                f"d={corrected_data.contact_displacement_um:.6g} um, "
                f"F={corrected_data.contact_load_mn:.6g} mN"
            )
            print()

        except Exception as exc:
            correction_failures += 1
            message = str(exc)
            print(
                f"[CONTACT FAILED] {xlsx_path.name}: {message}",
                file=sys.stderr,
            )
            summary_rows.append(
                {
                    "Filename": xlsx_path.name,
                    "Status": "contact detection failed",
                    "Worksheet": raw_data.sheet_name,
                    "Message": message,
                }
            )

    write_detection_summary(
        corrected_output_root / "contact_detection_summary.csv",
        summary_rows,
    )

    print(
        f"First two output stages completed: "
        f"{len(xlsx_files) - raw_failures} raw datasets and "
        f"{len(xlsx_files) - correction_failures} contact-corrected datasets "
        f"generated."
    )

    # Stage 3: only now request/consume approximate endpoint displacement and
    # generate the final endpoint-trimmed outputs.
    endpoint_summary_rows: list[dict[str, object]] = []
    endpoint_failures = 0
    endpoint_generated = 0

    if corrected_datasets:
        endpoint_output_root.mkdir(parents=True, exist_ok=True)

        print()
        print("Endpoint-trimmed output stage")
        print(
            "Provide an approximate displacement on the contact-corrected "
            "axis. The program searches after peak load for the nearby local "
            "load minimum."
        )

        for xlsx_path, corrected_data in corrected_datasets.items():
            if args.end_displacement_um is not None:
                approximate_displacement_um = args.end_displacement_um
            elif args.no_endpoint_prompt:
                approximate_displacement_um = None
            else:
                approximate_displacement_um = (
                    prompt_for_endpoint_displacement(
                        xlsx_path.name,
                        corrected_data,
                    )
                )

            if approximate_displacement_um is None:
                print(f"[ENDPOINT SKIPPED] {xlsx_path.name}")
                endpoint_summary_rows.append(
                    {
                        "Filename": xlsx_path.name,
                        "Status": "skipped",
                        "Message": (
                            "No approximate endpoint displacement was provided."
                        ),
                    }
                )
                continue

            try:
                endpoint_detection = detect_experiment_endpoint(
                    corrected_data,
                    approximate_displacement_um=(
                        approximate_displacement_um
                    ),
                    search_window_um=args.endpoint_search_window_um,
                    smooth_window_s=args.endpoint_smooth_window_s,
                )
                endpoint_data = apply_endpoint_trim(
                    corrected_data,
                    endpoint_detection,
                )

                endpoint_output_directory = (
                    endpoint_output_root / xlsx_path.stem
                )

                write_three_column_csv(
                    endpoint_output_directory
                    / f"{xlsx_path.stem}_processed_endpoint_trimmed.csv",
                    headers=(
                        "Time_from_contact_s",
                        "Displacement_from_contact_um",
                        "Load_from_contact_mN",
                    ),
                    time_s=endpoint_data.time_s,
                    displacement_um=endpoint_data.displacement_um,
                    load_mn=endpoint_data.load_mn,
                )
                save_standard_plot_set(
                    output_directory=endpoint_output_directory,
                    filename_stem=xlsx_path.stem,
                    output_suffix="endpoint_trimmed",
                    time_s=endpoint_data.time_s,
                    displacement_um=endpoint_data.displacement_um,
                    load_mn=endpoint_data.load_mn,
                    png_dpi=args.png_dpi,
                )
                save_endpoint_diagnostic(
                    corrected_data,
                    endpoint_detection,
                    output_stem=(
                        endpoint_output_directory
                        / f"{xlsx_path.stem}_endpoint_selection_diagnostic"
                    ),
                    png_dpi=args.png_dpi,
                )

                peak_index = endpoint_detection.peak_index
                endpoint_generated += 1

                endpoint_summary_rows.append(
                    {
                        "Filename": xlsx_path.name,
                        "Status": "endpoint selected",
                        "Approximate_displacement_um": format_number(
                            approximate_displacement_um
                        ),
                        "Search_window_um": format_number(
                            args.endpoint_search_window_um
                        ),
                        "Peak_time_s": format_number(
                            corrected_data.time_s[peak_index]
                        ),
                        "Peak_displacement_um": format_number(
                            corrected_data.displacement_um[peak_index]
                        ),
                        "Peak_load_mN": format_number(
                            corrected_data.load_mn[peak_index]
                        ),
                        "Selected_endpoint_time_s": format_number(
                            endpoint_data.endpoint_time_s
                        ),
                        "Selected_endpoint_displacement_um": format_number(
                            endpoint_data.endpoint_displacement_um
                        ),
                        "Selected_endpoint_load_mN": format_number(
                            endpoint_data.endpoint_load_mn
                        ),
                        "Retained_rows": endpoint_data.time_s.size,
                        "Message": "",
                    }
                )

                print(
                    f"[ENDPOINT OK] {xlsx_path.name}: "
                    f"requested d≈{approximate_displacement_um:.6g} um; "
                    f"selected local minimum at "
                    f"t={endpoint_data.endpoint_time_s:.6g} s, "
                    f"d={endpoint_data.endpoint_displacement_um:.6g} um, "
                    f"F={endpoint_data.endpoint_load_mn:.6g} mN"
                )

            except Exception as exc:
                endpoint_failures += 1
                message = str(exc)
                print(
                    f"[ENDPOINT FAILED] {xlsx_path.name}: {message}",
                    file=sys.stderr,
                )
                endpoint_summary_rows.append(
                    {
                        "Filename": xlsx_path.name,
                        "Status": "endpoint detection failed",
                        "Approximate_displacement_um": format_number(
                            approximate_displacement_um
                        ),
                        "Search_window_um": format_number(
                            args.endpoint_search_window_um
                        ),
                        "Message": message,
                    }
                )

        write_endpoint_summary(
            endpoint_output_root / "endpoint_selection_summary.csv",
            endpoint_summary_rows,
        )

    print()
    print(
        f"Completed: {len(xlsx_files) - raw_failures} raw, "
        f"{len(xlsx_files) - correction_failures} contact-corrected, and "
        f"{endpoint_generated} endpoint-trimmed datasets generated."
    )

    return 1 if raw_failures or correction_failures or endpoint_failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
